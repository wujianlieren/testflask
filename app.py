from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook
import pymysql, db

app = Flask(__name__)
def insert_data_to_mysql(data):
    conn = db.get_conn()
    try:
        cursor = conn.cursor()
        print("Data to be inserted:", data)
        # 假设表结构为 (id, name, age)，注意根据你的实际表结构修改
        cursor.executemany("INSERT INTO crm (company,tel,name,note,important,source) VALUES (%s, %s,%s,%s,%s,%s)", data)
        conn.commit()
    finally:
        conn.close()

@app.route('/upload_excel', methods=['GET', 'POST'])
def upload_excel():
    if request.method == 'POST':
        file = request.files['file']
        if file and file.filename.endswith('.xlsx'):
            try:
                workbook = load_workbook(file)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # 假设 Excel 的列顺序为 (name, age)，注意根据你的实际列顺序修改
                    data.append((row[0], row[1],row[2],row[3],row[4],row[5]))
                insert_data_to_mysql(data)
                return 'inset data success'
                #return redirect(url_for('show_data'))
            except Exception as e:
                return f'Error: {str(e)}'
        else:
            return 'Invalid file format. Please upload a valid Excel file.'
    return render_template('upload_excel.html')

#show data
@app.route('/show_data', methods=['GET', 'POST'])
def show_data():
    conn = db.get_conn()
    try:
        cursor=conn.cursor()
        sql='select * from crm'
        cursor.execute(sql)
        data=cursor.fetchall()
    except Exception as e:
        return f"Error {str(e)}"
    finally:
        conn.close()
    return render_template("show_data.html",data=data)


if __name__ == "__main__":
    app.run(debug=True)
